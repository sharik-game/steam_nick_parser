from openpyxl import Workbook
import tqdm

import config

def create_excel(raw_users: list[list[list[config.User]]], path: str):
  workbook = Workbook()
  sheet = workbook.active
  row = 1
  for c in tqdm.tqdm(raw_users, desc="Saving to excel"):
    for j in c:
      for k in j:
        sheet.cell(row=row, column=1, value=k.nickname) # type: ignore
        sheet.cell(row=row, column=2, value=k.description) # type: ignore
        sheet.cell(row=row, column=3, value=k.location) # type: ignore
        sheet.cell(row=row, column=4, value=k.name) # type: ignore
        sheet.cell(row=row, column=5, value=k.link) # type: ignore
        for i, other_name in enumerate(k.other_nicknames, start=6):
          if other_name == []: sheet.cell(row=row, column=i, value="") # type: ignore
          else: sheet.cell(row=row, column=i, value=other_name) # type: ignore
        row+=1
  workbook.save(path)
  if config.DEBUG >= 2: print("Saving excel was successful")
  if config.DEBUG >= 3: print(f"Excel file contains {row} rows")
